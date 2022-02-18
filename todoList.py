from datetime import datetime
import openpyxl as xl
from os import path

class toDoList:
    def __init__(self,work=None,time=None):
        self.work = work
        self.time = time
        if not path.exists("todoList.xlsx"):
            list = xl.Workbook()
            toDo = list.create_sheet("List",1)
            list.save("todoList.xlsx")
            list.close()
        self.list = xl.load_workbook("todoList.xlsx")
        self.toDo = self.list.get_sheet_by_name("List")

    def createNewToDo(self):
        self.toDo.append([self.work, self.time])
        self.list.save("todoList.xlsx")
        self.list.close()

    def showToDo(self):
        i = 1
        while i < 100:
            if self.toDo["A"+str(i)].value == None :
                break
            else:
                print(f"Yapılacak İş: {self.toDo['A'+str(i)].value} ---> Eklenme Zamanı: {self.toDo['B'+str(i)].value}")
                i += 1
        self.list.close()

while True:
    option = input("1-) Yapılacak İş Ekle\n2-) İşleri Göster\n0-) Çıkış\nSeçenek: ")
    if option == "1":
        work = input("Yapılacak İş: ")
        now = datetime.now()
        todo = toDoList(work,now)
        todo.createNewToDo()
    elif option == "2":
        todo = toDoList()
        todo.showToDo()
    elif option == "0":
        break
    else:
        print("Yanlış Seçim!")
